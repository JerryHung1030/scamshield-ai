import asyncio
import os
import sys
from dotenv import load_dotenv
import logging
import json
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
# 確保可以 import 同層或上層資料夾的模組
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))

# 以下為原先 managers / adapters / services import
from managers.embedding_manager import EmbeddingManager
from managers.vector_store_manager import VectorStoreManager
from managers.llm_manager import LLMManager
from managers.blacklist_manager import BlacklistManager
from managers.regulations_manager import RegulationsManager

from adapters.openai_adapter import OpenAIAdapter
from adapters.local_llama_adapter import LocalLlamaAdapter

from services.fraud_rag_service import FraudRAGService
from services.fraud_rag_service_reverse import FraudRAGService_reverse
from services.compliance_rag_service import ComplianceRAGService, LawComplianceService
from evaluation_utils import save_results_to_json, build_matrix, visualize_matrix

# 載入 .env
load_dotenv(override=True)

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

def find_big_label_id(label: str, fraud_labels: list) -> int:
    """
    根據 label 找到對應的 big_label_id。
    假設 label 格式為 "big_label:description"。
    """
    # 提取 big_label
    big_label = label.split(":", 1)[0]  # 只取冒號前的部分

    # 在 fraud_labels 中查找對應的 big_label_id
    for label_info in fraud_labels:
        if label_info["big_label"] == big_label:
            return label_info["big_label_id"]

    # 如果找不到，返回一個預設值或引發錯誤
    return -1 
def setup_managers():
    """根據 .env 配置，初始化所有核心 manager。"""

    # 1) 從 .env 讀取環境變數
    openai_api_key = os.getenv("OPENAI_API_KEY")
    qdrant_url = os.getenv("QDRANT_URL")
    qdrant_collection = os.getenv("QDRANT_COLLECTION")
    embedding_model_name = os.getenv("EMBED_MODEL")

    # 2) 建立 EmbeddingManager
    embedding_manager = EmbeddingManager(
        openai_api_key=openai_api_key,
        embedding_model_name=embedding_model_name
    )

    # 3) 建立 VectorStoreManager (Qdrant)
    vector_store_manager = VectorStoreManager(
        embedding_manager=embedding_manager,
        qdrant_url=qdrant_url,
        collection_name=qdrant_collection
    )

    # 4) 建立 LLMManager，並註冊多個 LLM Adapter
    llm_manager = LLMManager()

    openai_adapter = OpenAIAdapter(
        openai_api_key=openai_api_key,
        model_name = "gpt-4o-mini",
        temperature=0.0,
        max_tokens=1024
    )

    local_llama_adapter = LocalLlamaAdapter(
        model_path="models/llama.bin",
        temperature=0.0,
        max_tokens=2048
    )

    llm_manager.register_adapter("openai", openai_adapter)
    llm_manager.register_adapter("llama", local_llama_adapter)
    llm_manager.set_default_adapter("openai")

    # 4) 其他
    blacklist_manager = BlacklistManager(blacklist_db=["badurl.com", "lineid123"])
    regulations_manager = RegulationsManager(regulations_db={"some_law": "Lorem ipsum..."})

    return embedding_manager, vector_store_manager, llm_manager, blacklist_manager, regulations_manager

def strip_code_fence(text: str) -> str:
    """
    移除可能的 Markdown code fence
    """
    txt = text.strip()
    if txt.startswith("```"):
        txt = txt.split("\n", 1)[-1].strip()
    if txt.endswith("```"):
        txt = txt.rsplit("```", 1)[0].strip()
    return txt
def load_jsonl_file(path: str):
    """小幫手：讀取 JSON lines 檔並回傳 list[dict]."""
    if not path or not os.path.exists(path):
        logger.warning(f"File not found: {path}")
        return []
    all_data = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            all_data.append(json.loads(line))
    return all_data
async def extract_evidence_for_label(llm_manager, text: str, big_label: str) -> dict:
    """
    呼叫 LLM，請它針對 `text` + `big_label`，找出一個最關鍵的句子 evidence，
    並告訴我們它在 text 裡的 start/end index。
    回傳格式:
    {
      "evidence": "xxxx",
      "start_idx": 10,
      "end_idx": 25
    }
    若 LLM 無法解析，回傳空。
    """
    # 1) 組 Prompt
    prompt = f"""
    你是一個資訊擷取器，目標是從文本中找出最能證明它屬於「{big_label}」的關鍵句子，並準確標註該句子在整段文本中的起始與結束 index（以字元計算，開頭為 0）。

    請依據以下標準作答：
    1. 「關鍵句子」是最能反映分類標籤「{big_label}」的句子，應具有明確且直接的代表性。
    2. 請提供該句子在文本中的「原始文字」（不得修改內容），以及它在整段文字中出現的字元位置範圍。
    3. `start_idx` 為該句子第一個字的 index，`end_idx` 為最後一個字的 index（含該字）。
    4. 請確保 index 是以 **整段文本的第一個字為 index 0** 進行計算。
    5. 如果找不到合適的句子，請輸出空陣列 `[]`。

    請僅輸出 JSON，格式如下：
    {{
    "evidence": "關鍵句子",
    "start_idx": 起始位置（數字）,
    "end_idx": 結束位置（數字）
    }}

    ---
    文本：
    {text}
    """


    # 2) 呼叫 LLM
    adapter = llm_manager.get_adapter("openai")  # 你可視情況用 openai/llama
    raw_answer = await adapter.async_generate_response(prompt)

    # 3) 嘗試 parse
    #    小心 LLM 可能輸出包含 ```json 之類或是沒有依照格式
    cleaned = strip_code_fence(raw_answer)
    print(cleaned)
    try:
        parsed = json.loads(cleaned)
        if isinstance(parsed, dict) and "evidence" in parsed:
            return {
                "evidence": parsed["evidence"],
                "start_idx": parsed.get("start_idx", -1),
                "end_idx": parsed.get("end_idx", -1)
            }
        else:
            # 若 LLM 輸出 [] 或其他結構
            return {}
    except:
        return {}
async def run_multi_label_classification(doc_id: str):
    """
    將 15 個「詐騙大類」分別做為 user_query，跑 RAG，
    再把每次回傳結果組合成預期的多標籤輸出格式。
    """
    ############################
    # 1) 先定義 15 種詐騙分類
    #    big_label_id 與 big_label 需依你實際的表定映射
    ############################
    fraud_labels = [
        {"big_label_id": 1, "big_label": "假投資詐騙"},
        {"big_label_id": 2, "big_label": "網路購物詐騙"},
        {"big_label_id": 3, "big_label": "假買家騙賣家詐騙"},
        {"big_label_id": 4, "big_label": "假交友(投資詐財)詐騙"},
        {"big_label_id": 5, "big_label": "假中獎通知詐騙"},
        {"big_label_id": 6, "big_label": "假交友(徵婚詐財)詐騙"},
        {"big_label_id": 7, "big_label": "假求職詐騙"},
        {"big_label_id": 8, "big_label": "假借銀行貸款詐騙"},
        {"big_label_id": 9, "big_label": "假檢警詐騙"},
        {"big_label_id": 10, "big_label": "假廣告詐騙"},
        {"big_label_id": 11, "big_label": "釣魚簡訊(惡意連結)詐騙"},
        {"big_label_id": 12, "big_label": "騙取金融帳戶(卡片)詐騙"},
        {"big_label_id": 13, "big_label": "色情應召詐財詐騙"},
        {"big_label_id": 14, "big_label": "虛擬遊戲詐騙"},
        {"big_label_id": 15, "big_label": "猜猜我是誰詐騙"}
    ]

    ############################
    # 2) 建立 (或取得) Managers 與 FraudRAGService
    #    通常可以直接呼叫你 main.py 中的 setup_managers()... 
    ############################
    embedding_manager, vector_store_manager, llm_manager, blacklist_manager, regulations_manager = setup_managers()

    # 建立或取得你先前的 FraudRAGService 實例
    fraud_service = FraudRAGService_reverse(
        embedding_manager=embedding_manager,
        vector_store_manager=vector_store_manager,
        llm_manager=llm_manager,
        blacklist_manager=blacklist_manager,
        domain_key="FRAUD",
        selected_llm_name="openai"  # 假設你預設使用 openai
    )


    scam_file = os.getenv("SCAM_PATTERNS_FILE_big_label")
    scam_patterns = load_jsonl_file(scam_file)
    print(scam_patterns)
    # 如果這個 JSONL 其結構是 chunk-based（例如 uid、chunk_index、text），
    # 你可以用 add_jsonl_documents 形式，而後搜尋時就會保有 uid。
    # 這裡依你的實際欄位 key 做對應:
    vector_store_manager.add_jsonl_documents(
        domain="FRAUD",
        json_lines=scam_patterns,
        text_key="text",           # 該 chunk 的文字段
        meta_keys=["uid"]  # 你想要存的 metadata
    )

    ############################
    # 3) 逐一將 15 個分類當 user_query，呼叫 RAG
    ############################
    all_predictions = []
    scam_file_input = os.getenv("INPUT_FILE")
    chunk_list = load_jsonl_file(scam_file_input)
    for chunk in chunk_list:

        # c_id = chunk["uid"]
        # d_id = chunk["doc_id"]
        if chunk["doc_id"] != doc_id:
            print("不是該doc_id")
            continue
        text = f"uid:{chunk["uid"]} | {chunk["text"]}"


        # 執行 RAG 
        raw_answer = await fraud_service.generate_answer(user_query=text)
        # 注意: 由於 FraudRAGService 實作時回傳的是實際的 Python list (它先 json.loads 後再 return)，
        #       因此 raw_answer 就是 list[dict] 形式。若你要統一看成字串，請注意再度 json.loads。
        print(raw_answer)
        if not raw_answer:
            # 若 RAG 產生空陣列，就代表無命中
            print("無命中")
            continue

        # RAG 預期輸出資料結構是 list[dict]，裡頭含 doc_id, confidence(0~1), evidence 等等。
        # 依照你的需求把它轉換成多標籤預測的格式:
        for item in raw_answer:
            # 取得 uid (如果在 metadata 或 item 其他欄位)
            # 視你的 post_process() / doc_id / metadata 來對應
            uid = item.get("uid", "")  # 這裡假設 doc_id 即 uid
            confidence_score = item.get("confidence", 1.0)  # or 轉成 high/medium/low 字串
            # 以下示範若你要用 "high","medium","low"
            if isinstance(confidence_score, float):
                if confidence_score > 0.8:
                    confidence_txt = "high"
                elif confidence_score > 0.4:
                    confidence_txt = "medium"
                else:
                    confidence_txt = "low"
            else:
                confidence_txt = str(confidence_score)

            big_label_id = find_big_label_id(item["label"], fraud_labels)
            # 彙整到最終格式
            record = {
                "uid": uid,
                "big_label_id": big_label_id,
                "big_label": item["label"],
                "confidence": confidence_txt
            }
            all_predictions.append(record)

    # 建立 uid -> text 的對照表，以便後續根據 uid 取得原文
    chunk_text_map = {}
    for item in chunk_list:
        c_id = item.get("uid")
        c_text = item.get("text")
        chunk_text_map[c_id] = c_text

    # 第二階段: 對每個 (chunk, label) 呼叫 LLM 查找 evidence
    for pred in all_predictions:
        uid = pred["uid"]
        big_label = pred["big_label"]
        chunk_text = chunk_text_map.get(uid, "")

        if not chunk_text:
            continue

        # 呼叫 extract_evidence_for_label
        evidence_info = await extract_evidence_for_label(llm_manager, chunk_text, big_label)
        
        if evidence_info:
            # 例如:
            # evidence_info = {
            #   "evidence": "這句是關鍵呀！",
            #   "start_idx": 15,
            #   "end_idx": 25
            # }
            pred["evidence"] = [ evidence_info["evidence"] ]  # 可能只有一句
            pred["evidence_index"] = [ evidence_info["start_idx"], evidence_info["end_idx"] ]
        else:
            # 若沒有解析到關鍵句，就存空或自己定義
            pred["evidence"] = []
            pred["evidence_index"] = [-1, -1]
    ############################
    # 4) 彙整最終輸出 JSON，例如：
    ############################
    final_output = {
        "doc_id": doc_id,
        "direction": "reverse",       
        "model": "gpt-4o-mini",            
        "rag_k": 15,                  
        "conf_threshold": "low",      
        "predictions": all_predictions
    }

    # 這裡列印或可存檔 json
    print(json.dumps(final_output, ensure_ascii=False, indent=2))
    # 視需求返回
    return final_output

async def main():
    # 1) 打開或新建 Excel
    excel_path = "fraud_evaluation_data_format 1.xlsx"
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # 2) 取得或新建 sheet
    sheet_name = "reverse_4omini_k15_low"
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]

    # 3) 設定欄寬、樣式等（依需要）
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 60

    # 假使 row=1 可能留做欄位標題
    current_row = 2  # 從第二列開始寫

    # 4) 迭代 post0001 ~ post0030
    for i in range(1, 31):
        doc_id = f"post00{i:02d}"   # 產生 post0001, post0002, ... post0030
        print(f"Processing {doc_id} ...")
        # 執行多標籤推論
        result_json = await run_multi_label_classification(doc_id)

        # 5) 把 result_json dump 成字串，以便寫入 Excel
        result_str = json.dumps(result_json, ensure_ascii=False, indent=2)
        start_row = current_row
        end_row = current_row + 2

        ws.merge_cells(
            start_row=start_row, 
            start_column=2,  # B
            end_row=end_row, 
            end_column=2
        )

        # 7) 寫入合併後的儲存格
        cell = ws.cell(row=start_row, column=2)
        cell.value = result_str
        # 設定文字換行
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 8) row 往下走1行
        current_row = end_row + 1

    # 9) 存檔
    wb.save(excel_path)
    print(f"Done! Results saved to {excel_path}.")

if __name__ == "__main__":
    asyncio.run(main())