import asyncio
import os
import sys
from dotenv import load_dotenv
import logging
import json
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
# 確保可以 import 同層或上層資料夾的模組
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))

# 以下為原先 managers / adapters / services import
from managers.embedding_manager import EmbeddingManager
from managers.vector_store_manager import VectorStoreManager
from managers.llm_manager import LLMManager
from managers.blacklist_manager import BlacklistManager
from managers.regulations_manager import RegulationsManager

from adapters.openai_adapter import OpenAIAdapter
from adapters.local_llama_adapter import LocalLlamaAdapter

from services.fraud_rag_service import FraudRAGService
from services.compliance_rag_service import ComplianceRAGService, LawComplianceService
from evaluation_utils import save_results_to_json, build_matrix, visualize_matrix

# 載入 .env
load_dotenv(override=True)

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)


def setup_managers():
    """根據 .env 配置，初始化所有核心 manager。"""

    # 1) 從 .env 讀取環境變數
    openai_api_key = os.getenv("OPENAI_API_KEY")
    qdrant_url = os.getenv("QDRANT_URL")
    qdrant_collection = os.getenv("QDRANT_COLLECTION")
    embedding_model_name = os.getenv("EMBED_MODEL")

    # 2) 建立 EmbeddingManager
    embedding_manager = EmbeddingManager(
        openai_api_key=openai_api_key,
        embedding_model_name=embedding_model_name
    )

    # 3) 建立 VectorStoreManager (Qdrant)
    vector_store_manager = VectorStoreManager(
        embedding_manager=embedding_manager,
        qdrant_url=qdrant_url,
        collection_name=qdrant_collection
    )

    # 4) 建立 LLMManager，並註冊多個 LLM Adapter
    llm_manager = LLMManager()

    openai_adapter = OpenAIAdapter(
        openai_api_key=openai_api_key,
        model_name = "gpt-4o-mini",
        temperature=0.0,
        max_tokens=1024
    )

    local_llama_adapter = LocalLlamaAdapter(
        model_path="models/llama.bin",
        temperature=0.0,
        max_tokens=2048
    )

    llm_manager.register_adapter("openai", openai_adapter)
    llm_manager.register_adapter("llama", local_llama_adapter)
    llm_manager.set_default_adapter("openai")

    # 4) 其他
    blacklist_manager = BlacklistManager(blacklist_db=["badurl.com", "lineid123"])
    regulations_manager = RegulationsManager(regulations_db={"some_law": "Lorem ipsum..."})

    return embedding_manager, vector_store_manager, llm_manager, blacklist_manager, regulations_manager

def strip_code_fence(text: str) -> str:
    """
    移除可能的 Markdown code fence
    """
    txt = text.strip()
    if txt.startswith("```"):
        txt = txt.split("\n", 1)[-1].strip()
    if txt.endswith("```"):
        txt = txt.rsplit("```", 1)[0].strip()
    return txt
def load_jsonl_file(path: str):
    """小幫手：讀取 JSON lines 檔並回傳 list[dict]."""
    if not path or not os.path.exists(path):
        logger.warning(f"File not found: {path}")
        return []
    all_data = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            all_data.append(json.loads(line))
    return all_data
async def extract_evidence_for_label(llm_manager, text: str, big_label: str) -> dict:
    """
    呼叫 LLM，請它針對 `text` + `big_label`，找出一個最關鍵的句子 evidence，
    並告訴我們它在 text 裡的 start/end index。
    回傳格式:
    {
      "evidence": "xxxx",
      "start_idx": 10,
      "end_idx": 25
    }
    若 LLM 無法解析，回傳空。
    """
    # 1) 組 Prompt
    prompt = f"""
    你是一個資訊擷取器，目標是從文本中找出最能證明它屬於「{big_label}」的關鍵句子，並準確標註該句子在整段文本中的起始與結束 index（以字元計算，開頭為 0）。

    請依據以下標準作答：
    1. 「關鍵句子」是最能反映分類標籤「{big_label}」的句子，應具有明確且直接的代表性。
    2. 請提供該句子在文本中的「原始文字」（不得修改內容），以及它在整段文字中出現的字元位置範圍。
    3. `start_idx` 為該句子第一個字的 index，`end_idx` 為最後一個字的 index（含該字）。
    4. 請確保 index 是以 **整段文本的第一個字為 index 0** 進行計算。
    5. 如果找不到合適的句子，請輸出空陣列 `[]`。

    請僅輸出 JSON，格式如下：
    {{
    "evidence": "關鍵句子",
    "start_idx": 起始位置（數字）,
    "end_idx": 結束位置（數字）
    }}

    ---
    文本：
    {text}
    """


    # 2) 呼叫 LLM
    adapter = llm_manager.get_adapter("openai")  # 你可視情況用 openai/llama
    raw_answer = await adapter.async_generate_response(prompt)

    # 3) 嘗試 parse
    #    小心 LLM 可能輸出包含 ```json 之類或是沒有依照格式
    cleaned = strip_code_fence(raw_answer)
    print(cleaned)
    try:
        parsed = json.loads(cleaned)
        if isinstance(parsed, dict) and "evidence" in parsed:
            return {
                "evidence": parsed["evidence"],
                "start_idx": parsed.get("start_idx", -1),
                "end_idx": parsed.get("end_idx", -1)
            }
        else:
            # 若 LLM 輸出 [] 或其他結構
            return {}
    except:
        return {}
async def run_multi_label_classification(doc_id: str):
    """
    接收 doc_id 作為參數，執行多標籤分類流程，最後回傳 JSON 結構。
    """
    ############################
    # 1) 先定義 15 種詐騙分類
    ############################
    fraud_labels = [
        {"big_label_id": 1, "big_label": "假投資詐騙"},
        {"big_label_id": 2, "big_label": "網路購物詐騙"},
        {"big_label_id": 3, "big_label": "假買家騙賣家詐騙"},
        {"big_label_id": 4, "big_label": "假交友(投資詐財)詐騙"},
        {"big_label_id": 5, "big_label": "假中獎通知詐騙"},
        {"big_label_id": 6, "big_label": "假交友(徵婚詐財)詐騙"},
        {"big_label_id": 7, "big_label": "假求職詐騙"},
        {"big_label_id": 8, "big_label": "假借銀行貸款詐騙"},
        {"big_label_id": 9, "big_label": "假檢警詐騙"},
        {"big_label_id": 10, "big_label": "假廣告詐騙"},
        {"big_label_id": 11, "big_label": "釣魚簡訊(惡意連結)詐騙"},
        {"big_label_id": 12, "big_label": "騙取金融帳戶(卡片)詐騙"},
        {"big_label_id": 13, "big_label": "色情應召詐財詐騙"},
        {"big_label_id": 14, "big_label": "虛擬遊戲詐騙"},
        {"big_label_id": 15, "big_label": "猜猜我是誰詐騙"}
    ]

    ############################
    # 2) 建立(或取得) Managers 與 FraudRAGService
    ############################
    embedding_manager, vector_store_manager, llm_manager, blacklist_manager, regulations_manager = setup_managers()

    fraud_service = FraudRAGService(
        embedding_manager=embedding_manager,
        vector_store_manager=vector_store_manager,
        llm_manager=llm_manager,
        blacklist_manager=blacklist_manager,
        domain_key="FRAUD",
        selected_llm_name="openai"
    )

    # 載入 INPUT_FILE (其中包含許多 chunk 可能來自不同 doc_id)
    scam_file = os.getenv("INPUT_FILE")
    scam_patterns = load_jsonl_file(scam_file)
    # 如你所示：針對 scam_patterns 我們先 add_jsonl_documents 到 domain="FRAUD"
    vector_store_manager.add_jsonl_documents(
        domain="FRAUD",
        json_lines=scam_patterns,
        text_key="text",
        meta_keys=["uid", "doc_id", "chunk_index"]
    )

    ############################
    # 3) 逐一將 15 個詐騙分類當 user_query(S) -> filters={"doc_id": doc_id}
    ############################
    all_predictions = []
    for label_info in fraud_labels:
        user_query = label_info["big_label"]
        # 調用 generate_answer()，並用 filters={"doc_id": doc_id} 只抓取該 doc_id
        raw_answer = await fraud_service.generate_answer(
            user_query=user_query,
            filters={"doc_id": doc_id}
        )
        print(raw_answer)
        # 解析回傳
        if not raw_answer:
            continue

        # raw_answer 是 list[dict]
        for item in raw_answer:
            uid = item.get("uid", "")
            confidence_score = item.get("confidence", 1.0)

            # 轉成高/中/低
            if isinstance(confidence_score, float):
                if confidence_score > 0.8:
                    confidence_txt = "high"
                elif confidence_score > 0.4:
                    confidence_txt = "medium"
                else:
                    confidence_txt = "low"
            else:
                confidence_txt = str(confidence_score)

            record = {
                "uid": uid,
                "big_label_id": label_info["big_label_id"],
                "big_label": label_info["big_label"],
                "confidence": confidence_txt
            }
            all_predictions.append(record)

    # 建立 uid -> text 的 map
    # （若要進一步用 extract_evidence_for_label）
    chunk_text_map = {}
    for item in scam_patterns:
        c_id = item.get("uid")
        c_text = item.get("text")
        chunk_text_map[c_id] = c_text

    # 如果需要依 label 再取 evidence，可以再執行 extract_evidence_for_label
    for pred in all_predictions:
        uid = pred["uid"]
        big_label = pred["big_label"]
        chunk_text = chunk_text_map.get(uid, "")

        if not chunk_text:
            continue

        evidence_info = await extract_evidence_for_label(llm_manager, chunk_text, big_label)
        if evidence_info:
            pred["evidence"] = [ evidence_info["evidence"] ]
            pred["evidence_index"] = [ evidence_info["start_idx"], evidence_info["end_idx"] ]
        else:
            pred["evidence"] = []
            pred["evidence_index"] = [-1, -1]

   
    final_output = {
        "doc_id": doc_id,
        "direction": "forward",
        "model": "gpt-4o-mini",
        "rag_k": 3,  
        "conf_threshold": "low",
        "predictions": all_predictions
    }

    return final_output


async def main():
    # 1) 打開或新建 Excel
    excel_path = "fraud_evaluation_data_format 1.xlsx"
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # 2) 取得或新建 sheet
    sheet_name = "forward_4omini_k3_low"
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]

    # 3) 設定欄寬、樣式等（依需要）
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 60

    # 假使 row=1 可能留做欄位標題
    current_row = 2  # 從第二列開始寫

    # 4) 迭代 post0001 ~ post0030
    for i in range(1, 31):
        doc_id = f"post00{i:02d}"   # 產生 post0001, post0002, ... post0030
        print(f"Processing {doc_id} ...")
        # 執行多標籤推論
        result_json = await run_multi_label_classification(doc_id)

        # 5) 把 result_json dump 成字串，以便寫入 Excel
        result_str = json.dumps(result_json, ensure_ascii=False, indent=2)
        start_row = current_row
        end_row = current_row + 2

        ws.merge_cells(
            start_row=start_row, 
            start_column=2,  # B
            end_row=end_row, 
            end_column=2
        )

        # 7) 寫入合併後的儲存格
        cell = ws.cell(row=start_row, column=2)
        cell.value = result_str
        # 設定文字換行
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 8) row 往下走1行
        current_row = end_row + 1

    # 9) 存檔
    wb.save(excel_path)
    print(f"Done! Results saved to {excel_path}.")
if __name__ == "__main__":
    asyncio.run(main())